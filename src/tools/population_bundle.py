from __future__ import annotations

import ast
import calendar
import hashlib
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import date
from enum import Enum
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

from config import settings


class PopulationSourceRole(str, Enum):
    OPENING_BALANCE = "opening_balance"
    CIVIL_STATUS = "civil_status"
    RESIDENCE_MOVEMENT = "residence_movement"


class PopulationExtractionError(ValueError):
    pass


class PopulationStandardizationError(ValueError):
    pass


def _load_population_bundle_config(config_path: Path | str | None = None) -> dict[str, Any]:
    path = Path(config_path or settings.POPULATION_BUNDLE_CONFIG_PATH)
    if not path.exists():
        raise PopulationExtractionError(f"Population bundle config not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)
    if not isinstance(config, dict):
        raise PopulationExtractionError("Population bundle configuration must be a YAML mapping.")
    return config


def _evaluate_formula(formula: str, values: dict[str, int]) -> int:
    def _eval(node: ast.AST) -> int:
        if isinstance(node, ast.BinOp):
            left = _eval(node.left)
            right = _eval(node.right)
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            if isinstance(node.op, ast.Div):
                return left // right
            raise PopulationStandardizationError(f"Unsupported arithmetic operator: {type(node.op).__name__}")
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, ast.USub):
            return -_eval(node.operand)
        if isinstance(node, ast.Name):
            if node.id not in values:
                raise PopulationStandardizationError(f"Unknown formula field: {node.id}")
            return values[node.id]
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return int(node.value)
        raise PopulationStandardizationError(f"Unsupported expression in formula: {ast.dump(node)}")

    tree = ast.parse(formula, mode="eval")
    return _eval(tree.body)


def _config_key(value: object) -> str:
    return _key(value)


@dataclass(frozen=True)
class ReportingPeriod:
    kind: str
    start: date
    end: date
    label: str


@dataclass(frozen=True)
class OrganizationMetadata:
    organization_id: str
    organization_name: str


@dataclass
class PopulationExtractedSource:
    role: PopulationSourceRole
    source_filename: str
    source_sha256: str
    reporting_period: ReportingPeriod
    organization: OrganizationMetadata
    classification: str
    values: dict[str, int]
    detail_counts: dict[str, int]
    detail_record_count: int
    extraction_warnings: list[str]


def _key(value: object) -> str:
    text = unicodedata.normalize("NFC", str(value or "")).strip().casefold()
    text = re.sub(r"\s+", " ", text).rstrip(":")
    return text


def _integer(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip().replace(" ", "")
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    return None


class PopulationWorkbookExtractor:
    """Deterministic extraction for a population bundle workbook."""

    def extract(self, source_path: Path, source_filename: str | None = None) -> PopulationExtractedSource:
        if source_path.suffix.lower() != ".xlsx":
            raise PopulationExtractionError("population bundle sources must be XLSX files")
        filename = source_filename or source_path.name
        cached = self._read_workbook(source_path, data_only=True)
        formulas = self._read_workbook(source_path, data_only=False)
        role = self._detect_role(formulas)
        period = self._extract_reporting_period(cached)
        organization = self._extract_organization(cached)
        all_text = "\n".join(
            str(value)
            for rows in cached.values()
            for row in rows
            for value in row
            if value is not None
        )
        classification = (
            "synthetic_test_data"
            if "SYNTHETIC_TEST_DATA" in all_text.upper()
            else "official_candidate"
        )
        checksum = hashlib.sha256(source_path.read_bytes()).hexdigest()
        summaries, summary_locations = self._summary_values(cached)
        detail_counts, detail_records, detail_sheet = self._detail_counts(role, cached)
        values: dict[str, int] = {}
        warnings: list[str] = []

        for field in ROLE_FIELDS[role]:
            summary_value = summaries.get(field)
            detail_value = detail_counts.get(field)
            if summary_value is not None and detail_value is not None:
                if summary_value != detail_value:
                    raise PopulationExtractionError(
                        f"summary/detail mismatch for {field}: {summary_value} != {detail_value}"
                    )
                value = summary_value
            elif detail_value is not None:
                value = detail_value
                warnings.append(
                    f"{field} derived from detail counts because summary value was absent."
                )
            elif summary_value is not None:
                value = summary_value
            else:
                raise PopulationExtractionError(
                    f"required population indicator is missing: {field}"
                )
            if value < 0:
                raise PopulationExtractionError(
                    f"population indicator must be non-negative: {field}"
                )
            values[field] = value

        return PopulationExtractedSource(
            role=role,
            source_filename=filename,
            source_sha256=checksum,
            reporting_period=period,
            organization=organization,
            classification=classification,
            values=values,
            detail_counts=detail_counts,
            detail_record_count=detail_records,
            extraction_warnings=warnings,
        )

    @staticmethod
    def _read_workbook(path: Path, *, data_only: bool) -> dict[str, list[list[Any]]]:
        workbook = load_workbook(path, data_only=data_only, read_only=True)
        try:
            return {
                sheet.title: [list(row) for row in sheet.iter_rows(values_only=True)]
                for sheet in workbook.worksheets
            }
        finally:
            workbook.close()

    @staticmethod
    def _detect_role(workbook: dict[str, list[list[Any]]]) -> PopulationSourceRole:
        codes = {
            str(row[0]).strip()
            for rows in workbook.values()
            for row in rows
            if row and isinstance(row[0], str)
        }
        matches = [role for role, fields in ROLE_FIELDS.items() if fields.issubset(codes)]
        if len(matches) != 1:
            raise PopulationExtractionError(
                "population source role could not be determined uniquely"
            )
        return matches[0]

    @staticmethod
    def _extract_reporting_period(workbook: dict[str, list[list[Any]]]) -> ReportingPeriod:
        candidates: list[str] = []
        for rows in workbook.values():
            for row in rows:
                for index, value in enumerate(row):
                    if _key(value) == "kỳ báo cáo" and index + 1 < len(row):
                        candidates.insert(0, str(row[index + 1]))
                    elif value is not None:
                        candidates.append(str(value))
        for text in candidates:
            match = re.search(r"tháng\s*(\d{1,2})\s*[/_-]\s*(\d{4})", text, re.I)
            if match:
                month, year = int(match.group(1)), int(match.group(2))
                return ReportingPeriod(
                    kind="month",
                    start=date(year, month, 1),
                    end=date(year, month, calendar.monthrange(year, month)[1]),
                    label=f"Tháng {month:02d}/{year}",
                )
        raise PopulationExtractionError(
            "reporting period could not be determined from workbook content"
        )

    @staticmethod
    def _extract_organization(workbook: dict[str, list[list[Any]]]) -> OrganizationMetadata:
        for rows in workbook.values():
            for row in rows:
                for index, value in enumerate(row):
                    if _key(value) == "đơn vị tổng hợp" and index + 1 < len(row):
                        name = str(row[index + 1] or "").strip()
                        if name:
                            return OrganizationMetadata(
                                organization_id=hashlib.sha1(name.encode("utf-8")).hexdigest(),
                                organization_name=name,
                            )
        raise PopulationExtractionError(
            "reporting organization could not be determined"
        )

    @staticmethod
    def _summary_values(workbook: dict[str, list[list[Any]]]) -> tuple[dict[str, int], dict[str, tuple[str, int]]]:
        values: dict[str, int] = {}
        locations: dict[str, tuple[str, int]] = {}
        canonical = set().union(*ROLE_FIELDS.values())
        for sheet_name, rows in workbook.items():
            for row_number, row in enumerate(rows, start=1):
                if not row or str(row[0] or "").strip() not in canonical:
                    continue
                value = _integer(row[2] if len(row) > 2 else None)
                if value is not None:
                    values[str(row[0]).strip()] = value
                    locations[str(row[0]).strip()] = (sheet_name, row_number)
        return values, locations

    def _detail_counts(self, role: PopulationSourceRole, workbook: dict[str, list[list[Any]]]) -> tuple[dict[str, int], int, str | None]:
        if role is PopulationSourceRole.OPENING_BALANCE:
            return self._opening_counts(workbook)
        if role is PopulationSourceRole.CIVIL_STATUS:
            return self._civil_counts(workbook)
        return self._movement_counts(workbook)

    @staticmethod
    def _find_header(
        workbook: dict[str, list[list[Any]]], required: set[str]
    ) -> tuple[str, list[list[Any]], int, dict[str, int]]:
        for sheet_name, rows in workbook.items():
            for index, row in enumerate(rows):
                headers = {
                    _key(value): column
                    for column, value in enumerate(row)
                    if value is not None
                }
                if required.issubset(headers):
                    return sheet_name, rows, index, headers
        raise PopulationExtractionError(
            "required detail table was not found"
        )

    def _opening_counts(self, workbook: dict[str, list[list[Any]]]) -> tuple[dict[str, int], int, str]:
        sheet, rows, header_index, headers = self._find_header(
            workbook, {"mã địa bàn", "tổng thường trú", "tổng tạm trú"}
        )
        permanent = temporary = records = 0
        for row in rows[header_index + 1 :]:
            identity = str(row[headers["mã địa bàn"]] or "").strip()
            if _key(identity) == "tổng cộng":
                break
            if not identity:
                continue
            permanent_value = _integer(row[headers["tổng thường trú"]])
            temporary_value = _integer(row[headers["tổng tạm trú"]])
            if permanent_value is None or temporary_value is None:
                raise PopulationExtractionError(
                    "opening-balance detail contains a non-integer value"
                )
            permanent += permanent_value
            temporary += temporary_value
            records += 1
        return {
            "population_opening": permanent,
            "temporary_opening": temporary,
        }, records, sheet

    def _civil_counts(self, workbook: dict[str, list[list[Any]]]) -> tuple[dict[str, int], int, str]:
        sheet, rows, header_index, headers = self._find_header(
            workbook, {"mã sự kiện", "loại sự kiện", "thuộc dân cư thường trú xã"}
        )
        event_counts: Counter[str] = Counter()
        local_counts: Counter[str] = Counter()
        records = 0
        for row in rows[header_index + 1 :]:
            if not str(row[headers["mã sự kiện"]] or "").strip():
                continue
            event = str(row[headers["loại sự kiện"]] or "").strip().upper()
            if event not in {"KHAI_SINH", "KHAI_TU"}:
                continue
            event_counts[event] += 1
            if _key(row[headers["thuộc dân cư thường trú xã"]]) == "có":
                local_counts[event] += 1
            records += 1
        return {
            "birth_registered": event_counts["KHAI_SINH"],
            "birth_local_resident": local_counts["KHAI_SINH"],
            "death_registered": event_counts["KHAI_TU"],
            "death_local_resident": local_counts["KHAI_TU"],
        }, records, sheet

    def _movement_counts(self, workbook: dict[str, list[list[Any]]]) -> tuple[dict[str, int], int, str]:
        sheet, rows, header_index, headers = self._find_header(
            workbook, {"mã biến động", "loại biến động"}
        )
        counts: Counter[str] = Counter()
        accepted = {"THUONG_TRU_DEN", "THUONG_TRU_DI", "TAM_TRU_MOI", "TAM_TRU_KET_THUC"}
        records = 0
        for row in rows[header_index + 1 :]:
            if not str(row[headers["mã biến động"]] or "").strip():
                continue
            movement = str(row[headers["loại biến động"]] or "").strip().upper()
            if movement not in accepted:
                continue
            counts[movement] += 1
            records += 1
        return {
            "permanent_in": counts["THUONG_TRU_DEN"],
            "permanent_out": counts["THUONG_TRU_DI"],
            "temporary_new": counts["TAM_TRU_MOI"],
            "temporary_removed": counts["TAM_TRU_KET_THUC"],
        }, records, sheet


class PopulationBundleStandardizer:
    """Merge and validate three deterministic population source roles."""

    def standardize(self, sources: list[PopulationExtractedSource]) -> dict[str, Any]:
        self._validate_bundle_identity(sources)
        merged: dict[str, int] = {}

        for source in sources:
            for field, value in source.values.items():
                if field in merged and merged[field] != value:
                    raise PopulationStandardizationError(
                        f"conflicting canonical values for {field}: {merged[field]} != {value}"
                    )
                merged[field] = value

        required_inputs = set().union(*ROLE_FIELDS.values())
        missing = sorted(required_inputs - set(merged))
        if missing:
            raise PopulationStandardizationError(
                f"population bundle is missing canonical input fields: {', '.join(missing)}"
            )

        merged["population_closing"] = (
            merged["population_opening"]
            + merged["birth_local_resident"]
            + merged["permanent_in"]
            - merged["death_local_resident"]
            - merged["permanent_out"]
        )
        merged["temporary_closing"] = (
            merged["temporary_opening"]
            + merged["temporary_new"]
            - merged["temporary_removed"]
        )

        return {
            "values": merged,
            "classification": (
                "synthetic_test_data"
                if all(source.classification == "synthetic_test_data" for source in sources)
                else "official_candidate"
            ),
            "reporting_period": sources[0].reporting_period,
            "organization": sources[0].organization,
        }
